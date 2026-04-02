import os

from flask import Blueprint, request, jsonify, current_app
from datetime import datetime
import smtplib, traceback, json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import send_file
import base64
import mimetypes
import io
from flask import g
from routes.admin_guard import require_admin
from routes.admin_audit import log_admin_action
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook


# Assuming db.py and get_db_connection are set up correctly
from backend.db import get_db_connection

MASTER_REQUIRED_FIELDS = {
    "Supplier Registration Details": [
        "company_name_english",
        "contact_person_name",
        "contact_person_email",
        "contact_person_mobile",
        "city",
        "country",
        "cr_number",
        "cr_expiry_date",
        "computer_card_number",
        "computer_card_expiry_date",
        "signing_authority_name",
        "sponsor_name",
        "trade_license_name",
        "vat_tax_number",
        "category",
        "brand_name",
        "address",
        "street",
        "zone",
        "area"
    ],

    "Bank Details": [
        "bank_name",
        "iban",
        "account_holder_name",
        "bank_branch",
        "swift_code"
    ],

    "Required Documents": [
        "upload_trade_license_copy",
        "upload_vat_certificates_copy",
        "upload_cr_company",
        "upload_computer_card_copy",
        "upload_bank_letter",
        "upload_company_logo",
        "certificates"
    ],

    "Branch Details": [
        "branch_name_english",
        "branch_manager_name",
        "contact_number",
        "email",
        "street",
        "zone",
        "building",
        "office_no",
        "city",
        "country",
        "branch_license"
    ],

    "Store Details": [
        "store_name_english",
        "contact_person_name",
        "contact_person_mobile",
        "email",
        "street",
        "zone",
        "building",
        "shop_no",
        "operating_hours",
        "city",
        "country",
        "store_type",
        "delivery_pickup_availability"
    ]
}

approval_bp = Blueprint("approval_bp", __name__)

SUPPLIER_FILE_MAP = {
    "tradeLicense": "upload_trade_license_copy",
    "vatCertificate": "upload_vat_certificates_copy",
    "crCopy": "upload_cr_company",
    "computerCardCopy": "upload_computer_card_copy",
}


UPLOAD_FIELDS = set(SUPPLIER_FILE_MAP.values())



# ------------------------------------------


def send_email(to_email, subject, html_content):
    try:
        mail_server = current_app.config["MAIL_SERVER"]
        mail_port = current_app.config["MAIL_PORT"]
        mail_user = current_app.config["MAIL_USERNAME"]
        mail_pass = current_app.config["MAIL_PASSWORD"]
        sender_cfg = current_app.config["MAIL_DEFAULT_SENDER"]

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject

        if isinstance(sender_cfg, tuple):
            name, email_addr = sender_cfg
            msg["From"] = f"{name} <{email_addr}>"
        else:
            msg["From"] = sender_cfg

        msg["To"] = to_email
        msg.attach(MIMEText(html_content, "html"))

        with smtplib.SMTP(mail_server, mail_port) as server:
            server.starttls()
            server.login(mail_user, mail_pass)
            server.sendmail(msg["From"], to_email, msg.as_string())

        print(f"✅ Email sent to {to_email}")
        return True

    except Exception as e:
        print("❌ Email send failed:", e)
        return False


# =============================================================
# ROUTES
# =============================================================
@approval_bp.route("/admins/ops", methods=["GET"])
@require_admin("APPROVE_SUPPLIERS")
def get_ops_admins():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
SELECT 
    a.admin_id,
    a.name,
    COUNT(s.supplier_id) AS pending
FROM admin_users a
LEFT JOIN supplier_registration s
  ON s.assigned_admin_id = a.admin_id
 AND s.approval_status ILIKE 'Pending'
WHERE a.role_id = 2        -- <<< ASSUMING 2 = OPS_ADMIN in admin_roles
GROUP BY a.admin_id, a.name
ORDER BY pending ASC;

    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify({
        "admins": [
            {
                "admin_id": r["admin_id"],
                "name": r["name"],
                "pending": r["pending"]
            }
            for r in rows
        ]
    })

# -------------------------------------------------------------
# 1️⃣ GET PENDING SUPPLIERS (unchanged)
# -------------------------------------------------------------
@approval_bp.route("/suppliers/pending", methods=["GET"])
@require_admin("APPROVE_SUPPLIERS")
def get_pending_suppliers():

    conn = get_db_connection()
    cur = conn.cursor()

    admin_id = g.admin["admin_id"]

    # FIXED HERE
    role = (g.admin.get("role") or "").upper()

    print("DEBUG ADMIN:", g.admin)
    print("DEBUG ROLE:", role)

    if role == "SUPER_ADMIN":

        cur.execute("""
            SELECT 
                s.supplier_id,
                s.company_name_english AS company_name_en,
                s.contact_person_name,
                s.contact_person_email,
                s.contact_person_mobile,
                s.approval_status,
                s.created_at,
                s.assigned_admin_id,
                COALESCE(a.name, '') AS assigned_admin_name

            FROM supplier_registration s

            LEFT JOIN admin_users a
                ON a.admin_id = s.assigned_admin_id

            WHERE s.approval_status NOT IN ('Approved', 'Rejected')

            ORDER BY s.created_at DESC
        """)

    elif role == "OPS_ADMIN":

        cur.execute("""
            SELECT 
                s.supplier_id,
                s.company_name_english AS company_name_en,
                s.contact_person_name,
                s.contact_person_email,
                s.contact_person_mobile,
                s.approval_status,
                s.created_at,
                s.assigned_admin_id,
                COALESCE(a.name, '') AS assigned_admin_name

            FROM supplier_registration s

            LEFT JOIN admin_users a
                ON a.admin_id = s.assigned_admin_id

            WHERE s.assigned_admin_id = %s
            AND s.approval_status NOT IN ('Approved', 'Rejected')

            ORDER BY s.created_at DESC
        """, (admin_id,))

    else:
        print("ROLE NOT MATCHED:", role)
        return jsonify({"items": []})

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify({
        "items": [dict(r) for r in rows]
    })



# -------------------------------------------------------------
# 2️⃣ GET SPECIFIC SUPPLIER DETAILS  (unchanged)
# -------------------------------------------------------------
@approval_bp.route("/supplier/<int:supplier_id>", methods=["GET"])
@require_admin("APPROVE_SUPPLIERS")
def get_supplier(supplier_id):
    

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # SELECT all required fields, including the newly added file field
        cur.execute("""
    SELECT 
        supplier_id,
        contact_person_name,
        company_name_english,
        company_name_arabic,
        contact_person_email,
        contact_person_mobile,
        country,
        city,
        approval_status,   -- ✅ ADD THIS
        upload_trade_license_copy,
        upload_vat_certificates_copy,
        upload_cr_company,
        upload_computer_card_copy

    FROM supplier_registration
    WHERE supplier_id = %s
""", (supplier_id,))


        row = cur.fetchone()

        if not row:
            return jsonify({"error": "not found"}), 404

        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="READ_ACCESS",
            entity_type="supplier_profile",
            entity_id=supplier_id,
            ip_address=request.remote_addr
        )

        

        def extract_file(raw):
            if not raw:
                return None
            try:
                # Assuming the file column stores a JSON string with a 'filename' key
                obj = json.loads(raw)
                return obj.get("filename")
            except:
                return None

        # CLEAN + MINIMAL JSON FOR APPROVAL UI
        supplier_clean = {
    "supplier_id": row["supplier_id"],
    "approval_status": row["approval_status"],   # ✅ ADD THIS

    "fullName": row["contact_person_name"],
    "companyName": row["company_name_english"] or row["company_name_arabic"],
    "email": row["contact_person_email"],
    "phoneNumber": row["contact_person_mobile"],
    "country": row["country"],
    "city": row["city"],

    "tradeLicense": extract_file(row["upload_trade_license_copy"]),
    "vatCertificate": extract_file(row["upload_vat_certificates_copy"]),
    "crCopy": extract_file(row["upload_cr_company"]),
    "computerCardCopy": extract_file(row["upload_computer_card_copy"]),
}



        return jsonify({"data": supplier_clean})

    except Exception as e:
        print("❌ get_supplier:", e)
        traceback.print_exc()
        return jsonify({"error": "server error"}), 500

    finally:
        if cur: cur.close()
        if conn: conn.close()


# -------------------------------------------------------------
# 3️⃣ SEARCH SUPPLIERS (unchanged)
# -------------------------------------------------------------
@approval_bp.route("/suppliers/search", methods=["GET"])
@require_admin("APPROVE_SUPPLIERS")
def get_search_suppliers():
    

    filter_by = request.args.get("by", "").lower()
    filter_value = request.args.get("value", "").strip()

    if not filter_by or not filter_value:
        return jsonify({"error": "filter criteria missing"}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Base query and parameters
        query = """
            SELECT 
    s.supplier_id,
    s.company_name_english AS company_name_en,
    s.contact_person_name,
    s.contact_person_email,
    s.approval_status,
    s.created_at,
    s.assigned_admin_id,
    a.name AS assigned_admin_name
    FROM supplier_registration s
    LEFT JOIN admin_users a
    ON s.assigned_admin_id = a.admin_id
    WHERE 1=1

        """
        params = []

        # Build dynamic query based on filter_by
        if filter_by == "id" and filter_value.isdigit():
            query += " AND supplier_id = %s"
            params.append(int(filter_value))
        elif filter_by == "name":
            query += " AND company_name_english ILIKE %s"
            params.append(f"%{filter_value}%")
        elif filter_by == "status":
            query += " AND approval_status ILIKE %s"
            params.append(f"%{filter_value}%")
        else:
             # If the filter criteria are invalid or type mismatch for 'id'
            return jsonify({"error": "Invalid search type or value"}), 400

        query += " ORDER BY created_at DESC"

        cur.execute(query, tuple(params))
        rows = cur.fetchall()

        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="READ_ACCESS",
            entity_type="supplier_search",
            entity_id=0,
            new_value={"by": filter_by, "value": filter_value},
            ip_address=request.remote_addr
        )

        return jsonify({
    "items": [
        {
            "supplier_id": r["supplier_id"],
            "company_name_en": r["company_name_en"],
            "contact_person_name": r["contact_person_name"],
            "contact_person_email": r["contact_person_email"],
            "approval_status": r["approval_status"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "assigned_admin_id": r["assigned_admin_id"],
            "assigned_admin_name": r["assigned_admin_name"]
        }
        for r in rows
    ]
})


    except Exception as e:
        print("❌ get_search_suppliers:", e)
        return jsonify({"error": "server error"}), 500

    finally:
        if cur: cur.close()
        if conn: conn.close()
        
# -------------------------------------------------------------
# 4️⃣ REVIEW SUPPLIER (APPROVE / REJECT / RESUBMIT)
# -------------------------------------------------------------
@approval_bp.route("/supplier/<int:supplier_id>/review", methods=["PATCH"])
@require_admin("APPROVE_SUPPLIERS")
def review_supplier(supplier_id):
    

    data = request.get_json() or {}
    action = (data.get("action") or "").lower()

    conn = cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT approval_status,
                contact_person_email,
                contact_person_name,
                company_name_english,
                contact_person_mobile
            FROM supplier_registration
            WHERE supplier_id = %s
            FOR UPDATE
        """, (supplier_id,))
        supplier = cur.fetchone()

        if not supplier:
            return jsonify({"error": "supplier not found"}), 404

        supplier_email = supplier["contact_person_email"]
        contact_name = supplier["contact_person_name"]
        company = supplier["company_name_english"]
        mobile = supplier["contact_person_mobile"]
        old_status = supplier["approval_status"]

        if old_status == "Approved":
            return jsonify({"error": "approved supplier cannot be modified"}), 409

        admin_role = (g.admin.get("role") or "").upper()

        if action == "approve" and admin_role != "SUPER_ADMIN":
            return jsonify({
                "error": "Only SUPER ADMIN can approve suppliers"
            }), 403

        # -------------------------------------------------
        # APPROVE
        # -------------------------------------------------
        if action == "approve":

            admin_role = (g.admin.get("role") or "").upper()

            # Only SUPER ADMIN allowed
            if admin_role != "SUPER_ADMIN":
                return jsonify({
                    "error": "Only SUPER ADMIN can approve"
                }), 403

            # Must be Under Review state
            if old_status != "Under Review":
                return jsonify({
                    "error": "Supplier must be Under Review before approval"
                }), 400

            cur.execute("""
                UPDATE supplier_registration
                SET approval_status = 'Approved',
                    approved_by_admin_id = %s,
                    approved_at = NOW(),
                    resubmit_reason = NULL,
                    resubmit_at = NULL,
                    updated_at = NOW()
                WHERE supplier_id = %s
            """, (
                g.admin["admin_id"],
                supplier_id
            ))


# -------------------------------------------
# CREATE / ACTIVATE USER (OTP LOGIN ENABLED)
# -------------------------------------------
            cur.execute("""
    INSERT INTO users (username, role, linked_id, is_first_login, status)
    VALUES (%s, 'supplier', %s, TRUE, 'active')
    ON CONFLICT (username)
    DO UPDATE SET
        role = EXCLUDED.role,
        linked_id = EXCLUDED.linked_id,
        is_first_login = TRUE,
        status = 'active'
""", (
    supplier_email,
    supplier_id
))


            conn.commit()
            log_admin_action(
                admin_id=g.admin["admin_id"],
                action="APPROVE_SUPPLIER",
                entity_type="supplier",
                entity_id=supplier_id,
                old_value={"approval_status": old_status},
                new_value={"approval_status": "Approved"},
                ip_address=request.remote_addr
            )

            login_url = os.getenv("FRONTEND_BASE_URL", "https://mahal-app") + "/SupplierLogIn"

            send_email(
                supplier_email,
                "Supplier Approved",
                f"""
                <html>
                <body style="font-family:Arial, sans-serif; background:#f4f6f9; padding:20px;">
                    <div style="max-width:600px; margin:auto; background:#ffffff;
                                padding:24px; border-radius:8px;">

                        <p>Dear <b>{contact_name}</b>,</p>

                        <p>Your supplier account for <b>{company}</b> has been approved.</p>



                        <p>
                            <b>Company:</b> {company}<br>
                            <b>Mobile:</b> {mobile}
                        </p>

                        <div style="text-align:center; margin-top:25px;">
                            <a href="{login_url}"
                               style="background:#ff8c00;
                                      color:#ffffff;
                                      padding:12px 26px;
                                      text-decoration:none;
                                      border-radius:6px;
                                      font-weight:bold;
                                      display:inline-block;">
                                Login to Supplier Dashboard
                            </a>
                        </div>



                    </div>
                </body>
                </html>
                """
            )

        # -------------------------------------------------
        # REJECT
        # -------------------------------------------------
        elif action == "reject":
            reason = data.get("reason", "").strip()
            if not reason:
                return jsonify({"error": "rejection reason required"}), 400

            cur.execute("""
                UPDATE supplier_registration
                SET approval_status = 'Rejected',
                    rejection_reason = %s,
                    updated_at = NOW()
                WHERE supplier_id = %s
            """, (reason, supplier_id))

            conn.commit()
            log_admin_action(
                admin_id=g.admin["admin_id"],
                action="REJECT_SUPPLIER",
                entity_type="supplier",
                entity_id=supplier_id,
                old_value={"approval_status": old_status},
                new_value={"approval_status": "Rejected", "reason": reason},
                ip_address=request.remote_addr
            )

            send_email(
                supplier_email,
                "Supplier Rejected",
                f"<p>{company} registration rejected.<br>Reason: {reason}</p>"
            )

        # -------------------------------------------------
        # RESUBMIT
        # -------------------------------------------------
        elif action == "resubmit":
            reason = data.get("reason", "").strip()
            if not reason:
                return jsonify({"error": "resubmission reason required"}), 400

            cur.execute("""
                UPDATE supplier_registration
                SET approval_status = 'Resubmit',
                    resubmit_reason = %s,
                    resubmit_at = NOW(),
                    updated_at = NOW()
                WHERE supplier_id = %s
            """, (reason, supplier_id))

            conn.commit()
            log_admin_action(
                admin_id=g.admin["admin_id"],
                action="RESUBMIT_SUPPLIER",
                entity_type="supplier",
                entity_id=supplier_id,
                old_value={"approval_status": old_status},
                new_value={"approval_status": "Resubmit", "reason": reason},
                ip_address=request.remote_addr
            )

            send_email(
                supplier_email,
                "Resubmission Required",
                f"<p>Additional documents required for {company}.<br>{reason}</p>"
            )

        else:
            return jsonify({"error": "invalid action"}), 400

        return jsonify({"message": f"Supplier {supplier_id} {action}ed successfully"})

    except Exception as e:
        print("❌ review_supplier:", e)
        traceback.print_exc()
        return jsonify({"error": "server error"}), 500

    finally:
        if cur: cur.close()
        if conn: conn.close()



@approval_bp.route("/supplier/<int:supplier_id>/file/<string:field_name>", methods=["GET"])
@require_admin("APPROVE_SUPPLIERS")
def admin_get_supplier_file(supplier_id, field_name):
    

    db_field = SUPPLIER_FILE_MAP.get(field_name)
    if not db_field:
        return jsonify({"error": "invalid file field"}), 400
    
    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="READ_ACCESS",
        entity_type="supplier_file",
        entity_id=supplier_id,
        new_value={"field": field_name},
        ip_address=request.remote_addr
    )

    conn = cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"SELECT {db_field} FROM supplier_registration WHERE supplier_id = %s", (supplier_id,))
        row = cur.fetchone()

        if not row or not row[db_field]:
            return jsonify({"error": "file not found"}), 404

        blob = row[db_field]

        # 1. Handle RAW BYTE STORAGE
        if isinstance(blob, (bytes, bytearray, memoryview)):
            if isinstance(blob, memoryview):
                blob = blob.tobytes()
            return send_file(io.BytesIO(blob), mimetype="application/pdf") # Default to PDF for old records

        # 2. Handle JSON + BASE64 STORAGE (New WebP/PDF records)
        data = json.loads(blob)
        file_bytes = base64.b64decode(data["content"])
        
        # Priority 1: Use stored mimetype. Priority 2: Guess by filename. Priority 3: Default.
        mimetype = (
            data.get("mimetype")
            or mimetypes.guess_type(data.get("filename", ""))[0]
            or "application/pdf"
        )

        # Special Fix for WebP detection on some systems
        filename = data.get("filename") or "file"

        if "webp" in filename.lower() and "webp" not in mimetype:
            mimetype = "image/webp"

        response = send_file(
            io.BytesIO(file_bytes),
            mimetype=mimetype,
            as_attachment=False, # This allows browser preview
            download_name=data.get("filename")
        )

        # 3. Add headers to ensure PDF/WebP displays inline
        response.headers["Content-Disposition"] = f"inline; filename={data.get('filename')}"
        # Prevent browsers from forcing a download
        response.headers["X-Content-Type-Options"] = "nosniff"
        
        return response

    except Exception as e:
        print("❌ Admin file serve error:", e)
        return jsonify({"error": "server error"}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()

def fetch_basic(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_registration"

    cur.execute(f"""
        SELECT
            contact_person_name,
            company_name_english,
            contact_person_email
        FROM {table}
        WHERE supplier_id = %s
    """, (entity_id,))

    row = cur.fetchone()
    cur.close(); conn.close()

    if not row:
        return {}

    return {
        "supplierName": row["contact_person_name"],
        "companyName": row["company_name_english"],
        "email": row["contact_person_email"]
    }

def fetch_org(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_registration" if role == "supplier" else "restaurant_registration"

    cur.execute(f"SELECT * FROM {table} WHERE {role}_id = %s", (entity_id,))
    row = cur.fetchone()

    cur.close(); conn.close()
    return dict(row) if row else {}

def fetch_address(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_registration" if role == "supplier" else "restaurant_registration"

    cur.execute(f"SELECT * FROM {table} WHERE {role}_id = %s", (entity_id,))
    row = cur.fetchone()

    cur.close(); conn.close()
    return dict(row) if row else {}

def fetch_bank(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_registration" if role == "supplier" else "restaurant_registration"

    cur.execute(f"SELECT * FROM {table} WHERE {role}_id = %s", (entity_id,))
    row = cur.fetchone()

    cur.close(); conn.close()
    return dict(row) if row else {}

def fetch_files(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_registration" if role == "supplier" else "restaurant_registration"

    cur.execute(f"""
        SELECT
            upload_trade_license_copy AS tradeLicenseCopy,
            upload_vat_certificates_copy AS vatCertificate,
            upload_cr_company AS crCopy,
            upload_computer_card_copy AS compCardCopy,
            upload_bank_letter AS bankLetter,
            upload_company_logo AS companyLogo,
            certificates AS certificates
        FROM {table}
        WHERE {role}_id = %s
    """, (entity_id,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return {}

    return {k: bool(v) for k, v in row.items()}

def fetch_branches(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_branch_registration" if role == "supplier" else "restaurant_branch_registration"

    cur.execute(f"SELECT * FROM {table} WHERE {role}_id = %s", (entity_id,))
    rows = cur.fetchall()

    cur.close(); conn.close()
    return [dict(r) for r in rows] if rows else []

def fetch_store(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_store_registration" if role == "supplier" else "restaurant_store_registration"

    cur.execute(f"""
        SELECT
            store_name_english,
            contact_person_name,
            contact_person_mobile,
            email,
            street,
            zone,
            building,
            shop_no,
            operating_hours,
            city,
            country,
            store_type,
            delivery_pickup_availability
        FROM {table}
        WHERE supplier_id = %s
    """, (entity_id,))

    # row = cur.fetchone()
    rows = cur.fetchall()

    cur.close()
    conn.close()

    # return dict(row) if row else {}
    return [dict(r) for r in rows] if rows else []

def get_filled_profile_data(role, entity_id):
    data = {}

    data.update(fetch_basic(role, entity_id))
    data.update(fetch_org(role, entity_id))
    data.update(fetch_address(role, entity_id))
    data.update(fetch_bank(role, entity_id))
    # data.update(fetch_store(role, entity_id))
    data.update(fetch_files(role, entity_id))

    data["branches"] = fetch_branches(role, entity_id)
    data["stores"] = fetch_store(role, entity_id)

    return data

def calculate_remaining_fields(filled):
    remaining = {}

    for section, fields in MASTER_REQUIRED_FIELDS.items():
        missing = []

        for field in fields:

            # if section == "Branch Details":
            if section in ["Branch Details", "Store Details"]:
                continue

            value = filled.get(field)

            if not value or str(value).strip() == "":
                missing.append(field)

        if missing:
            remaining[section] = missing

    # -------- Branches --------
    branches = filled.get("branches", [])
    branch_missing = []

    if not branches:
        for f in MASTER_REQUIRED_FIELDS["Branch Details"]:
            branch_missing.append(f)

    else:
        for i, b in enumerate(branches, start=1):
            for f in MASTER_REQUIRED_FIELDS["Branch Details"]:
                if not b.get(f):
                    branch_missing.append(f"Branch {i} - {f}")

    if branch_missing:
        remaining["Branch Details"] = branch_missing

    # -------- Stores --------
    stores = filled.get("stores", [])
    store_missing = []

    if not stores:
        # 🔥 If no stores at all, list ALL store fields
        for f in MASTER_REQUIRED_FIELDS["Store Details"]:
            store_missing.append(f)
    else:
        for i, s in enumerate(stores, 1):
            for f in MASTER_REQUIRED_FIELDS["Store Details"]:
                if not s.get(f):
                    store_missing.append(f"Store {i} - {f}")

    if store_missing:
        remaining["Store Details"] = store_missing

    # ✅ ONLY return dictionary
    return remaining


def pretty(field):
    return (
        field.replace("En", " (English)")
             .replace("_", " ")
             .title()
    )

def generate_corporate_mail(company, supplier_name, remaining):
    sections_html = ""

    for section, fields in remaining.items():
        items = "".join(
            f"<li>{pretty(f)}</li>" for f in fields
        )
        sections_html += f"""
        <h4 style="margin-bottom:6px;">{section}</h4>
        <ul>{items}</ul>
        """

    return f"""
    <html>
    <body style="font-family:Arial; background:#f4f6f9; padding:20px;">
      <div style="max-width:650px; background:#fff; padding:24px; border-radius:8px;">

        <p>
          Dear <strong>{supplier_name}</strong>,<br/>
          <span style="color:#555;">({company})</span>
        </p>

        <p>
          During profile verification, we found that the following
          information/documents are still pending.
        </p>

        {sections_html}

        <p>
          Kindly reply to this email with the above details.
          Our team will update your profile accordingly.
        </p>

        <p>
          Regards,<br/>
          <strong>Onboarding & Compliance Team</strong>
        </p>

      </div>
    </body>
    </html>
    """

@approval_bp.route("/send-remaining-profile-mail", methods=["POST"])
@require_admin("APPROVE_SUPPLIERS")
def send_remaining_profile_mail():

    

    role = request.json.get("role")
    entity_id = request.json.get("id")

    if role != "supplier":
        return jsonify({"error": "invalid role"}), 400

    if not entity_id:
        return jsonify({"error": "invalid supplier id"}), 400
    
    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="SEND_REMAINING_PROFILE_MAIL",
        entity_type=role,
        entity_id=entity_id,
        ip_address=request.remote_addr
    )

    # supplier = fetch_basic(role, entity_id)
    filled = get_filled_profile_data(role, entity_id)

    remaining = calculate_remaining_fields(filled)

    supplier = fetch_basic(role, entity_id)

    company_name = supplier.get("companyName") or "Supplier"
    supplier_email = supplier.get("email")

    if not entity_id:
        return jsonify({
            "status": False,
            "message": "Invalid supplier id"
        }), 400

    if not remaining:
        return jsonify({"status": False, "message": "Profile already complete"}), 400

    if not supplier_email:
        return jsonify({
            "status": False,
            "message": "Supplier email not found"
        }), 400

    html = generate_corporate_mail(
        company_name,
        supplier.get("supplierName", "Supplier"),
        remaining
    )

    send_email(
        supplier_email,
        "Action Required: Remaining Profile Information",
        html
    )

    return jsonify({"status": True})

def fetch_email(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor()

    table = "supplier_registration" if role == "supplier" else "restaurant_registration"
    id_col = "supplier_id" if role == "supplier" else "restaurant_id"

    cur.execute(
        f"SELECT contact_person_email FROM {table} WHERE {id_col} = %s",
        (entity_id,)
    )

    row = cur.fetchone()
    cur.close()
    conn.close()

    return row["contact_person_email"] if row else None

from openpyxl import Workbook

def build_supplier_excel():
    wb = Workbook()

    # =====================================================
    # SHEET 1: SUPPLIER REGISTRATION
    # =====================================================
    ws = wb.active
    ws.title = "Supplier Registration"

    ws.append([
        "company_name_english",
        "company_name_arabic",
        "company_email",
        "contact_person_name",
        "contact_person_mobile",
        "contact_person_email",
        "address",
        "street",
        "zone",
        "area",
        "city",
        "country",
        "category",
        "brand_name",
        "cr_number",
        "cr_expiry_date",
        "computer_card_number",
        "computer_card_expiry_date",
        "signing_authority_name",
        "sponsor_name",
        "trade_license_name",
        "vat_tax_number",
        "bank_name",
        "bank_branch",
        "iban",
        "account_holder_name",
        "swift_code"
    ])

    # =====================================================
    # SHEET 2: SUPPLIER BRANCH DETAILS
    # =====================================================
    branch = wb.create_sheet("Supplier Branches")
    branch.append([
        "branch_name_english",
        "branch_name_arabic",
        "branch_manager_name",
        "contact_number",
        "email",
        "street",
        "zone",
        "building",
        "office_no",
        "city",
        "country",
        "branch_license"
    ])

    # =====================================================
    # SHEET 3: SUPPLIER STORE DETAILS
    # =====================================================
    store = wb.create_sheet("Supplier Stores")
    store.append([
        "branch_name",
        "store_name_english",
        "store_name_arabic",
        "contact_person_name",
        "contact_person_mobile",
        "email",
        "street",
        "zone",
        "building",
        "shop_no",
        "operating_hours",
        "store_type",
        "delivery_pickup_availability",
        "city",
        "country"
    ])

    # =====================================================
    # SHEET 4: DOCUMENTS (INSTRUCTION ONLY)
    # =====================================================
    docs = wb.create_sheet("Documents")
    docs.append(["document_field", "instruction"])

    documents = [
        "upload_cr_company",
        "upload_computer_card_copy",
        "upload_trade_license_copy",
        "upload_vat_certificates_copy",
        "upload_bank_letter",
        "upload_company_logo",
        "certificates"
    ]

    for d in documents:
        docs.append([d, "Attach this document in email"])

    return wb


def send_excel_mail(to_email, workbook, filename):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    msg = MIMEMultipart()
    msg["Subject"] = "Action Required: Profile Completion Excel Template"

    sender_cfg = current_app.config["MAIL_DEFAULT_SENDER"]
    if isinstance(sender_cfg, tuple):
        name, email_addr = sender_cfg
        msg["From"] = f"{name} <{email_addr}>"
        from_addr = email_addr
    else:
        msg["From"] = sender_cfg
        from_addr = sender_cfg

    msg["To"] = to_email

    msg.attach(MIMEText(
        "Please fill the attached Excel template and attach required documents in reply to this email.",
        "plain"
    ))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(buffer.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={filename}")
    msg.attach(part)

    server = smtplib.SMTP(
        current_app.config["MAIL_SERVER"],
        current_app.config["MAIL_PORT"],
        timeout=20
    )

    try:
        server.connect(
            current_app.config["MAIL_SERVER"],
            current_app.config["MAIL_PORT"]
        )
        server.ehlo()
        server.starttls()
        server.ehlo()

        server.login(
            current_app.config["MAIL_USERNAME"],
            current_app.config["MAIL_PASSWORD"]
        )

        server.sendmail(from_addr, [to_email], msg.as_string())

    finally:
        server.quit()

@approval_bp.route("/suppliers/send-excel-template", methods=["POST"])
@require_admin("APPROVE_SUPPLIERS")
def send_excel_template():
    

    data = request.get_json()
    role = data.get("role")
    entity_id = data.get("id")

    if role != "supplier":
        return jsonify({"error": "invalid role"}), 400

    if not entity_id:
        return jsonify({"error": "invalid supplier id"}), 400

    email = fetch_email("supplier", entity_id)
    if not email:
        return jsonify({"error": "email not found"}), 400

    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="SEND_EXCEL_TEMPLATE",
        entity_type="supplier",
        entity_id=entity_id,
        ip_address=request.remote_addr
    )

    workbook = build_supplier_excel()
    filename = "supplier_profile_template.xlsx"

    try:
        send_excel_mail(email, workbook, filename)
    except Exception as e:
        print("❌ Excel mail failed:", e)
        return jsonify({"error": "failed to send email"}), 500

    return jsonify({"status": True, "message": "Supplier Excel template sent"})


@approval_bp.route("/supplier/<int:supplier_id>/assign", methods=["PATCH"])
@require_admin("APPROVE_SUPPLIERS")
def assign_supplier(supplier_id):

    if (g.admin.get("role") or "").upper() != "SUPER_ADMIN":
        return jsonify({
            "error": "Only SUPER ADMIN can assign suppliers"
        }), 403

    data = request.get_json() or {}
    admin_id = data.get("admin_id")  # can be None

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔥 LOCK ROW
    cur.execute("""
        SELECT assigned_admin_id
        FROM supplier_registration
        WHERE supplier_id = %s
        FOR UPDATE
    """, (supplier_id,))
    
    row = cur.fetchone()

    if not row:
        return jsonify({"error": "supplier not found"}), 404

    old_admin = row["assigned_admin_id"]

    # 🔥 MAIN FIX
    if admin_id:
        # ASSIGN
        cur.execute("""
            UPDATE supplier_registration
            SET assigned_admin_id = %s,
                approval_status = 'Assigned',
                updated_at = NOW()
            WHERE supplier_id = %s
        """, (admin_id, supplier_id))

    else:
        # 🔥 DEASSIGN FIX
        cur.execute("""
            UPDATE supplier_registration
            SET assigned_admin_id = NULL,
                approval_status = 'Pending',
                updated_at = NOW()
            WHERE supplier_id = %s
        """, (supplier_id,))

    conn.commit()

    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="ASSIGN_SUPPLIER",
        entity_type="supplier",
        entity_id=supplier_id,
        old_value={"assigned_admin_id": old_admin},
        new_value={"assigned_admin_id": admin_id},
        ip_address=request.remote_addr
    )

    cur.close()
    conn.close()

    return jsonify({"status": True})

# =========================================================
# AUTO ASSIGN SUPPLIERS (LEAST PENDING OPS ADMIN FIRST)
# =========================================================
@approval_bp.route("/suppliers/auto-assign", methods=["PATCH"])
@require_admin("MANAGE_ADMIN_USERS")
def auto_assign_suppliers():

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # 1️⃣ GET OPS ADMINS ordered by least pending suppliers
        cur.execute("""
            SELECT 
                a.admin_id,

                COUNT(s.supplier_id) FILTER (
                    WHERE s.approval_status ILIKE 'Pending'
                ) AS pending

            FROM admin_users a

            JOIN admin_roles ar
                ON ar.role_id = a.role_id

            LEFT JOIN supplier_registration s
                ON s.assigned_admin_id = a.admin_id
               AND s.approval_status ILIKE 'Pending'

            WHERE ar.role_name = 'OPS_ADMIN'

            GROUP BY a.admin_id

            ORDER BY pending ASC, a.admin_id ASC
        """)

        admins = cur.fetchall()

        if not admins:
            return jsonify({"error": "No OPS admins found"}), 400

        admin_ids = [a["admin_id"] for a in admins]

        # 2️⃣ GET UNASSIGNED SUPPLIERS
        cur.execute("""
            SELECT supplier_id
            FROM supplier_registration
            WHERE approval_status ILIKE 'Pending'
              AND assigned_admin_id IS NULL
            ORDER BY created_at ASC
        """)

        suppliers = cur.fetchall()

        if not suppliers:
            return jsonify({
                "status": True,
                "message": "No unassigned suppliers"
            })

        # 3️⃣ ROUND ROBIN ASSIGNMENT
        index = 0

        for s in suppliers:

            admin_id = admin_ids[index]

            cur.execute("""
    UPDATE supplier_registration
    SET assigned_admin_id = %s,
        approval_status = 'Assigned',
        updated_at = NOW()
    WHERE supplier_id = %s
""", (admin_id, s["supplier_id"]))



            index = (index + 1) % len(admin_ids)

        conn.commit()

        # 4️⃣ AUDIT LOG
        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="AUTO_ASSIGN_SUPPLIERS",
            entity_type="supplier",
            entity_id=0,
            ip_address=request.remote_addr
        )

        return jsonify({
            "status": True,
            "message": "Auto assignment completed"
        })

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ auto_assign_suppliers:", e)
        traceback.print_exc()

        return jsonify({"error": "server error"}), 500

    finally:

        if cur: cur.close()
        if conn: conn.close()
@approval_bp.route("/supplier/<int:supplier_id>/complete-profile", methods=["PATCH"])
@require_admin("APPROVE_SUPPLIERS")
def complete_profile(supplier_id):

    admin_role = (g.admin.get("role") or "").upper()

    if admin_role not in ["OPS_ADMIN", "SUPPORT_ADMIN"]:
        return jsonify({
            "error": f"Only OPS or SUPPORT can complete profile. Your role: {admin_role}"
        }), 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    UPDATE supplier_registration
    SET approval_status = 'Profile Completed',
        profile_completed_at = NOW(),
        updated_at = NOW()
    WHERE supplier_id = %s
    AND assigned_admin_id = %s
    AND approval_status = 'Assigned'
    """, (
    supplier_id,
    g.admin["admin_id"]
    ))

    if cur.rowcount == 0:
        conn.rollback()
        return jsonify({
            "error": "Supplier not assigned to you OR invalid current status"
        }), 400

    conn.commit()


    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="PROFILE_COMPLETED",
        entity_type="supplier",
        entity_id=supplier_id
    )

    cur.close()
    conn.close()

    return jsonify({"status": True})
@approval_bp.route("/supplier/<int:supplier_id>/send-to-review", methods=["PATCH"])
@require_admin("APPROVE_SUPPLIERS")
def send_to_review(supplier_id):

    admin_role = (g.admin.get("role") or "").upper()

    if admin_role not in ["OPS_ADMIN", "SUPPORT_ADMIN"]:
        return jsonify({
            "error": f"Only OPS or SUPPORT can send for review. Your role: {admin_role}"
        }), 403
    
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    UPDATE supplier_registration
    SET approval_status = 'Under Review',
        sent_to_review_at = NOW(),
        reviewed_by_admin_id = %s,
        reviewed_at = NOW(),
        updated_at = NOW()
    WHERE supplier_id = %s
    AND assigned_admin_id = %s
    AND approval_status = 'Profile Completed'
    """, (
    g.admin["admin_id"],
    supplier_id,
    g.admin["admin_id"]
    ))

    if cur.rowcount == 0:
        conn.rollback()
        return jsonify({
            "error": "Supplier must be Profile Completed and assigned to you"
        }), 400

    conn.commit()


    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="SENT_TO_SUPER_ADMIN",
        entity_type="supplier",
        entity_id=supplier_id
    )

    cur.close()
    conn.close()

    return jsonify({"status": True})
