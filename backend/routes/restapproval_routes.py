
from flask import Blueprint, request, jsonify, current_app, send_file
import smtplib, traceback, json, base64, mimetypes, io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from routes.admin_guard import require_admin
from routes.admin_audit import log_admin_action
from flask import g
from email import encoders
from openpyxl import Workbook
from email.mime.base import MIMEBase
from psycopg2.extras import RealDictCursor



from backend.db import get_db_connection

RESTAURANT_REQUIRED_FIELDS = {

    "Restaurant Registration Details": [
        "restaurant_name_english",
        "cr_number",
        "cr_expiry_date",
        "computer_card_number",
        "computer_card_expiry_date",
        "signing_authority_name",
        "sponsor_name",
        "trade_license_name",
        "vat_tax_number",
        "contact_person_name",
        "contact_person_mobile",
        "contact_person_email",
        "restaurant_email_address",
        "address",
        "street",
        "zone",
        "area",
        "city",
        "country"
    ],

    "Bank Details": [
        "bank_name",
        "iban",
        "account_holder_name",
        "bank_branch",
        "swift_code"
    ],

    "Required Documents": [
        "upload_cr_copy",
        "upload_computer_card_copy",
        "upload_trade_license_copy",
        "upload_vat_certificate_copy",
        "upload_food_safety_certificate",
        "upload_company_logo"
    ],

    "Branch Details": [
        "branch_name_english",
        "branch_manager_name",
        "contact_number",
        "email",
        "street",
        "zone",
        "building",
        "office_number",
        "city",
        "country"
    ],

    "Store Details": [
        "store_name_english",
        "contact_person_name",
        "contact_person_mobile",
        "email",
        "street",
        "zone",
        "building",
        "shop_no",
        "operating_hours",
        "city",
        "country"
    ]
}


# =========================================================
# BLUEPRINT
# =========================================================
restapproval_bp = Blueprint(
    "restapproval_bp",
    __name__,
    url_prefix="/api/v1/admin"
)


# =========================================================
# FILE MAP (RESTAURANT ONLY)
# =========================================================
RESTAURANT_FILE_MAP = {
    "tradeLicense": "upload_trade_license_copy",
    "vatCertificate": "upload_vat_certificate_copy",
    "foodSafetyCertificate": "upload_food_safety_certificate",
}

def send_email(to_email, subject, html_content):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject

        sender_cfg = current_app.config["MAIL_DEFAULT_SENDER"]
        msg["From"] = (
            f"{sender_cfg[0]} <{sender_cfg[1]}>"
            if isinstance(sender_cfg, tuple)
            else sender_cfg
        )

        msg["To"] = to_email
        msg.attach(MIMEText(html_content, "html"))

        with smtplib.SMTP(
            current_app.config["MAIL_SERVER"],
            current_app.config["MAIL_PORT"]
        ) as server:
            server.starttls()
            server.login(
                current_app.config["MAIL_USERNAME"],
                current_app.config["MAIL_PASSWORD"],
            )
            server.sendmail(msg["From"], to_email, msg.as_string())

        return True
    except Exception as e:
        print("❌ Email error:", e)
        return False


def extract_filename(raw):
    if not raw:
        return None
    try:
        return json.loads(raw).get("filename")
    except Exception:
        return None


# =========================================================
# 1️⃣ GET PENDING RESTAURANTS
# =========================================================
@restapproval_bp.route("/restaurants/pending", methods=["GET"])
@require_admin("APPROVE_RESTAURANTS")
def get_pending_restaurants():
    conn = cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("""
            SELECT 
    r.restaurant_id,
    r.restaurant_name_english AS restaurant_name_en,
    r.contact_person_name,
    r.contact_person_email,
    r.contact_person_mobile,
    r.approval_status,
    r.created_at,
    r.assigned_admin_id,
    a.name AS assigned_admin_name

FROM restaurant_registration r

LEFT JOIN admin_users a
    ON r.assigned_admin_id = a.admin_id

WHERE r.approval_status = 'Pending'

ORDER BY r.created_at DESC

        """)

        rows = cur.fetchall()

        # ✅ AUDIT — READ
        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="READ_ACCESS",
            entity_type="restaurant_list_pending",
            entity_id=0,
            ip_address=request.remote_addr
        )

        return jsonify({
            "items": [
                {
                    "restaurant_id": r["restaurant_id"],
                    "restaurant_name_en": r["restaurant_name_en"],
                    "contact_person_name": r["contact_person_name"],
                    "contact_person_email": r["contact_person_email"],
                    "contact_person_mobile": r["contact_person_mobile"],
                    "approval_status": r["approval_status"],
                    "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                    "assigned_admin_id": r["assigned_admin_id"],
                    "assigned_admin_name": r["assigned_admin_name"],

                }
                for r in rows
            ]
        })

    except Exception as e:
        print("❌ get_pending_restaurants:", e)
        return jsonify({"error": "server error"}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()


# =========================================================
# 2️⃣ GET RESTAURANT DETAILS
# =========================================================
@restapproval_bp.route("/restaurant/<int:restaurant_id>", methods=["GET"])
@require_admin("APPROVE_RESTAURANTS")
def get_restaurant(restaurant_id):



    conn = cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("""
            SELECT restaurant_id,
                   restaurant_name_english,
                   restaurant_name_arabic,
                   contact_person_name,
                   contact_person_email,
                   contact_person_mobile,
                   country,
                   city,
                   upload_trade_license_copy,
                   upload_vat_certificate_copy,
                   upload_food_safety_certificate
            FROM restaurant_registration
            WHERE restaurant_id = %s
        """, (restaurant_id,))

        row = cur.fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        
        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="READ_ACCESS",
            entity_type="restaurant_profile",
            entity_id=restaurant_id,
            ip_address=request.remote_addr
        )
        

        return jsonify({
            "data": {
                "restaurant_id": row["restaurant_id"],
                "restaurantName": row["restaurant_name_english"] or row["restaurant_name_arabic"],
                "fullName": row["contact_person_name"],
                "email": row["contact_person_email"],
                "phoneNumber": row["contact_person_mobile"],
                "country": row["country"],
                "city": row["city"],
                "tradeLicense": extract_filename(row["upload_trade_license_copy"]),
                "vatCertificate": extract_filename(row["upload_vat_certificate_copy"]),
                "foodSafetyCertificate": extract_filename(row["upload_food_safety_certificate"]),
            }
        })

    except Exception as e:
        if conn:
            conn.rollback()

        print("❌ review_restaurant:", e)
        traceback.print_exc()
        return jsonify({"error": "server error"}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()



# =========================================================
# 3️⃣ SEARCH RESTAURANTS
# =========================================================
@restapproval_bp.route("/restaurants/search", methods=["GET"])
@require_admin("APPROVE_RESTAURANTS")
def search_restaurants():


    by = request.args.get("by", "").lower()
    value = request.args.get("value", "").strip()

    if not by or not value:
        return jsonify({"error": "filter missing"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        query = """
            SELECT 
    r.restaurant_id,
    r.restaurant_name_english AS restaurant_name_en,
    r.contact_person_name,
    r.contact_person_email,
    r.approval_status,
    r.created_at,
    r.assigned_admin_id,
    a.name AS assigned_admin_name

FROM restaurant_registration r

LEFT JOIN admin_users a
    ON r.assigned_admin_id = a.admin_id

WHERE 1=1

        """
        params = []

        if by == "id" and value.isdigit():
            query += " AND restaurant_id = %s"
            params.append(int(value))
        elif by == "name":
            query += " AND restaurant_name_english ILIKE %s"
            params.append(f"%{value}%")
        elif by == "status":
            query += " AND approval_status ILIKE %s"
            params.append(f"%{value}%")
        else:
            return jsonify({"error": "invalid filter"}), 400

        query += " ORDER BY created_at DESC"
        cur.execute(query, tuple(params))
        rows = cur.fetchall()

        log_admin_action(
    admin_id=g.admin["admin_id"],
    action="READ_ACCESS",
    entity_type="restaurant_search",
    entity_id=0,
    new_value={"by": by, "value": value},
    ip_address=request.remote_addr
)

        return jsonify({
            "items": [
                {
                    "restaurant_id": r["restaurant_id"],
                    "restaurant_name_en": r["restaurant_name_en"],
                    "contact_person_name": r["contact_person_name"],
                    "contact_person_email": r["contact_person_email"],
                    "approval_status": r["approval_status"],
                    "created_at": r["created_at"].isoformat() if r["created_at"] else None,
"assigned_admin_id": r["assigned_admin_id"],
"assigned_admin_name": r["assigned_admin_name"],

                }
                for r in rows
            ]
        })

    except Exception as e:
        if conn:
            conn.rollback()
        print("❌ search_restaurants:", e)
        return jsonify({"error": "server error"}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()



# =========================================================
# 4️⃣ REVIEW RESTAURANT (APPROVE / REJECT / RESUBMIT)
# =========================================================
@restapproval_bp.route("/restaurant/<int:restaurant_id>/review", methods=["PATCH"])
@require_admin("APPROVE_RESTAURANTS")
def review_restaurant(restaurant_id):


    data = request.get_json() or {}
    action = (data.get("action") or "").lower()

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("""
            SELECT approval_status,
                contact_person_email,
                contact_person_name,
                restaurant_name_english
            FROM restaurant_registration
            WHERE restaurant_id = %s
            FOR UPDATE
        """, (restaurant_id,))
        r = cur.fetchone()

        if not r:
            return jsonify({"error": "restaurant not found"}), 404

        email = r["contact_person_email"]
        name = r["contact_person_name"]
        restaurant = r["restaurant_name_english"]
        old_status = r["approval_status"] or "UNKNOWN"
        if old_status == "Approved":
            return jsonify({"error": "restaurant already approved"}), 409


        # -------------------------------------------------
        # APPROVE
        # -------------------------------------------------
        if action == "approve":
            

            cur.execute("""
                UPDATE restaurant_registration
                SET approval_status='Approved',
                    rejection_reason=NULL,
                    resubmit_reason=NULL,
                    resubmit_at=NULL,
                    updated_at=NOW()
                WHERE restaurant_id=%s
            """, (restaurant_id,))

            cur.execute("""
                INSERT INTO users (username, role, linked_id, is_first_login, status)
                VALUES (%s, 'restaurant', %s, TRUE, 'active')
                ON CONFLICT (username)
                DO UPDATE SET
                    role = EXCLUDED.role,
                    linked_id = EXCLUDED.linked_id,
                    is_first_login = TRUE,
                    status = 'active'
            """, (email, restaurant_id))


            conn.commit()

            log_admin_action(
                admin_id=g.admin["admin_id"],
                action="APPROVE_RESTAURANT",
                entity_type="restaurant",
                entity_id=restaurant_id,
                old_value={"approval_status": old_status},
                new_value={"approval_status": "Approved"},
                ip_address=request.remote_addr
            )

            login_url = "http://localhost:3000/RestaurantLogIn"

            send_email(
                email,
                "Restaurant Approved",
                f"""
                <html>
                <body style="font-family:Arial, sans-serif; background:#f4f6f9; padding:20px;">
                    <div style="max-width:600px; margin:auto; background:#ffffff;
                                padding:24px; border-radius:8px;">

                        <p>Dear <b>{name}</b>,</p>

                        <p>Your restaurant <b>{restaurant}</b> has been approved.</p>

                        <p>
                            <b>Username:</b> {email}<br>
                            
                        </p>

                        <div style="text-align:center; margin-top:25px;">
                            <a href="{login_url}"
                               style="background:#ff8c00;
                                      color:#ffffff;
                                      padding:12px 26px;
                                      text-decoration:none;
                                      border-radius:6px;
                                      font-weight:bold;
                                      display:inline-block;">
                                Login to Dashboard
                            </a>
                        </div>
                    </div>
                </body>
                </html>
                """
            )

        # -------------------------------------------------
        # REJECT
        # -------------------------------------------------
        elif action == "reject":
            reason = data.get("reason", "").strip()
            if not reason:
                return jsonify({"error": "reason required"}), 400

            cur.execute("""
                UPDATE restaurant_registration
                SET approval_status='Rejected',
                    rejection_reason=%s,
                    updated_at=NOW()
                WHERE restaurant_id=%s
            """, (reason, restaurant_id))
            conn.commit()

            log_admin_action(
                admin_id=g.admin["admin_id"],
                action="REJECT_RESTAURANT",
                entity_type="restaurant",
                entity_id=restaurant_id,
                old_value={"approval_status": old_status},
                new_value={"approval_status": "Rejected", "reason": reason},
                ip_address=request.remote_addr
            )

            send_email(
                email,
                "Restaurant Rejected",
                f"<p>{restaurant} rejected.<br>Reason: {reason}</p>"
            )

        # -------------------------------------------------
        # RESUBMIT
        # -------------------------------------------------
        elif action == "resubmit":
            reason = data.get("reason", "").strip()
            if not reason:
                return jsonify({"error": "reason required"}), 400

            cur.execute("""
                UPDATE restaurant_registration
                SET approval_status='Resubmit',
                    resubmit_reason=%s,
                    resubmit_at=NOW(),
                    updated_at=NOW()
                WHERE restaurant_id=%s
            """, (reason, restaurant_id))
            conn.commit()
            log_admin_action(
                admin_id=g.admin["admin_id"],
                action="RESUBMIT_RESTAURANT",
                entity_type="restaurant",
                entity_id=restaurant_id,
                old_value={"approval_status": old_status},
                new_value={"approval_status": "Resubmit", "reason": reason},
                ip_address=request.remote_addr
            )

            send_email(
                email,
                "Resubmission Required",
                f"<p>Resubmission needed for {restaurant}.<br>{reason}</p>"
            )

        else:
            return jsonify({"error": "invalid action"}), 400

        return jsonify({"message": "Restaurant review updated"})

    except Exception as e:
        if conn:
            conn.rollback()

        print("❌ review_restaurant:", e)
        traceback.print_exc()
        return jsonify({"error": "server error"}), 500

    finally:
        cur.close()
        conn.close()

# =========================================================
# 5️⃣ ADMIN FILE PREVIEW
# =========================================================
@restapproval_bp.route("/restaurant/<int:restaurant_id>/file/<string:field_name>", methods=["GET"])
@require_admin("APPROVE_RESTAURANTS")
def admin_get_restaurant_file(restaurant_id, field_name):

    db_field = RESTAURANT_FILE_MAP.get(field_name)
    if not db_field:
        return jsonify({"error": "invalid file field"}), 400

    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="READ_ACCESS",
        entity_type="restaurant_file",
        entity_id=restaurant_id,
        new_value={
            "field": field_name,
            "db_column": db_field
        },
        ip_address=request.remote_addr
    )


    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute(
            f"SELECT {db_field} FROM restaurant_registration WHERE restaurant_id=%s",
            (restaurant_id,)
        )
        row = cur.fetchone()

        if not row or not row[db_field]:
            return jsonify({"error": "file not found"}), 404

        blob = row[db_field]

        if isinstance(blob, (bytes, bytearray, memoryview)):
            if isinstance(blob, memoryview):
                blob = blob.tobytes()
            return send_file(
                io.BytesIO(blob),
                mimetype="application/pdf",
                as_attachment=False
            )

        data = json.loads(blob)
        file_bytes = base64.b64decode(data["content"])
        mimetype = (
            data.get("mimetype")
            or mimetypes.guess_type(data.get("filename", ""))[0]
            or "application/pdf"
        )

        return send_file(
            io.BytesIO(file_bytes),
            mimetype=mimetype,
            as_attachment=False,
            download_name=data["filename"],
        )

    except Exception as e:
        print("❌ admin_get_restaurant_file:", e)
        return jsonify({"error": "server error"}), 500
    finally:
        cur.close()
        conn.close()

def fetch_restaurant_registration(restaurant_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT *
        FROM restaurant_registration
        WHERE restaurant_id = %s
    """, (restaurant_id,))

    row = cur.fetchone()
    cur.close(); conn.close()

    return dict(row) if row else {}

def fetch_restaurant_branches(restaurant_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT *
        FROM restaurant_branch_registration
        WHERE restaurant_id = %s
    """, (restaurant_id,))

    rows = cur.fetchall()
    cur.close(); conn.close()

    return [dict(r) for r in rows] if rows else []

def fetch_restaurant_stores(restaurant_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT
            store_name_english,
            contact_person_name,
            contact_person_mobile,
            email,
            street,
            zone,
            building,
            shop_no,
            operating_hours,
            city,
            country
        FROM restaurant_store_registration
        WHERE restaurant_id = %s
    """, (restaurant_id,))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [dict(r) for r in rows] if rows else []

def fetch_restaurant_documents(restaurant_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT
            upload_cr_copy,
            upload_computer_card_copy,
            upload_trade_license_copy,
            upload_vat_certificate_copy,
            upload_food_safety_certificate,
            upload_company_logo
        FROM restaurant_registration
        WHERE restaurant_id = %s
    """, (restaurant_id,))

    row = cur.fetchone()
    cur.close(); conn.close()

    if not row:
        return {}

    return {k: bool(v) for k, v in row.items()}

def get_full_restaurant_profile(restaurant_id):
    data = {}

    data.update(fetch_restaurant_registration(restaurant_id))
    data.update(fetch_restaurant_documents(restaurant_id))

    data["branches"] = fetch_restaurant_branches(restaurant_id)
    data["stores"] = fetch_restaurant_stores(restaurant_id)

    return data

def calculate_restaurant_remaining_fields(filled):
    remaining = {}

    # ---------- MAIN SECTIONS ----------
    for section, fields in RESTAURANT_REQUIRED_FIELDS.items():
        if section in ["Branch Details", "Store Details"]:
            continue

        missing = []
        for field in fields:
            val = filled.get(field)
            if val is None or str(val).strip() == "":
                missing.append(field)

        if missing:
            remaining[section] = missing

    # ---------- BRANCHES ----------
    branches = filled.get("branches", [])
    branch_missing = []

    if not branches:
        # ✅ NO branches → list ALL branch fields
        for f in RESTAURANT_REQUIRED_FIELDS["Branch Details"]:
            branch_missing.append(f)
    else:
        for i, b in enumerate(branches, 1):
            for f in RESTAURANT_REQUIRED_FIELDS["Branch Details"]:
                if not b.get(f):
                    branch_missing.append(f"Branch {i} - {f}")

    if branch_missing:
        remaining["Branch Details"] = branch_missing

    # ---------- STORES ----------
    stores = filled.get("stores", [])
    store_missing = []

    if not stores:
        # ✅ NO stores → list ALL store fields
        for f in RESTAURANT_REQUIRED_FIELDS["Store Details"]:
            store_missing.append(f)
    else:
        for i, s in enumerate(stores, 1):
            for f in RESTAURANT_REQUIRED_FIELDS["Store Details"]:
                if not s.get(f):
                    store_missing.append(f"Store {i} - {f}")

    if store_missing:
        remaining["Store Details"] = store_missing

    return remaining

def pretty(field):
    return field.replace("_", " ").title()

def generate_restaurant_remaining_mail(restaurant, remaining):
    sections = ""

    for section, fields in remaining.items():
        items = "".join(f"<li>{pretty(f)}</li>" for f in fields)
        sections += f"<h4>{section}</h4><ul>{items}</ul>"

    return f"""
    <html>
    <body style="font-family:Arial">
        <p>Dear <b>{restaurant}</b>,</p>

        <p>
        During profile verification, we found the following mandatory
        information/documents are missing.
        </p>

        {sections}

        <p>
        Kindly reply to this email with the above details.
        Our team will update your profile accordingly.
        </p>

        <p>Regards,<br/>Onboarding Team</p>
    </body>
    </html>
    """

@restapproval_bp.route("/restaurants/send-remaining-profile-mail", methods=["POST"])
@require_admin("APPROVE_RESTAURANTS")
def send_remaining_restaurant_profile_mail():



    restaurant_id = request.json.get("id")

    if not restaurant_id:
        return jsonify({
        "status": False,
        "message": "Invalid restaurant id"
    }), 400

    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="SEND_REMAINING_PROFILE_MAIL",
        entity_type="restaurant",
        entity_id=restaurant_id,
        ip_address=request.remote_addr
    )

    # 1️⃣ Fetch profile
    filled = get_full_restaurant_profile(restaurant_id)

    # 2️⃣ Calculate missing fields
    remaining = calculate_restaurant_remaining_fields(filled)

    if not remaining:
        return jsonify({
            "status": False,
            "message": "Profile already complete"
        }), 400

    # 3️⃣ Fetch basic info
    restaurant = fetch_restaurant_registration(restaurant_id)
    email = restaurant.get("contact_person_email")
    name = restaurant.get("restaurant_name_english") or "Restaurant"

    if not email:
        return jsonify({
            "status": False,
            "message": "Restaurant email not found"
        }), 400

    # 4️⃣ Generate email
    html = generate_restaurant_remaining_mail(name, remaining)

    # 5️⃣ Send email
    sent = send_email(
        email,
        "Action Required: Remaining Profile Information",
        html
    )

    if not sent:
        return jsonify({
            "status": False,
            "message": "Failed to send email"
        }), 500

    return jsonify({"status": True})

def fetch_email(role, entity_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if role == "supplier":
        cur.execute("""
            SELECT contact_person_email
            FROM supplier_registration
            WHERE supplier_id = %s
        """, (entity_id,))
        row = cur.fetchone()
        email = row["contact_person_email"] if row else None

    else:  # restaurant
        cur.execute("""
            SELECT
                COALESCE(contact_person_email, restaurant_email_address) AS email
            FROM restaurant_registration
            WHERE restaurant_id = %s
        """, (entity_id,))
        row = cur.fetchone()
        email = row["email"] if row else None

    cur.close()
    conn.close()
    return email

def build_restaurant_excel():
    wb = Workbook()

    # =====================================================
    # SHEET 1: RESTAURANT REGISTRATION
    # =====================================================
    ws = wb.active
    ws.title = "Restaurant Registration"

    ws.append([
        "restaurant_name_english",
        "restaurant_name_arabic",
        "restaurant_email_address",
        "contact_person_name",
        "contact_person_mobile",
        "contact_person_email",
        "address",
        "street",
        "zone",
        "area",
        "city",
        "country",
        "cr_number",
        "cr_expiry_date",
        "computer_card_number",
        "computer_card_expiry_date",
        "signing_authority_name",
        "sponsor_name",
        "trade_license_name",
        "vat_tax_number",
        "bank_name",
        "bank_branch",
        "iban",
        "account_holder_name",
        "swift_code"
    ])

    # =====================================================
    # SHEET 2: RESTAURANT BRANCH DETAILS
    # =====================================================
    branch = wb.create_sheet("Restaurant Branches")
    branch.append([
        "branch_name_english",
        "branch_name_arabic",
        "branch_manager_name",
        "contact_number",
        "email",
        "street",
        "zone",
        "building",
        "office_number",
        "city",
        "country"
    ])

    # =====================================================
    # SHEET 3: RESTAURANT STORE DETAILS
    # =====================================================
    store = wb.create_sheet("Restaurant Stores")
    store.append([
        "branch_name",
        "store_name_english",
        "store_name_arabic",
        "contact_person_name",
        "contact_person_mobile",
        "email",
        "street",
        "zone",
        "building",
        "shop_no",
        "operating_hours",
        "city",
        "country"
    ])

    # =====================================================
    # SHEET 4: DOCUMENTS
    # =====================================================
    docs = wb.create_sheet("Documents")
    docs.append(["document_field", "instruction"])

    documents = [
        "upload_cr_copy",
        "upload_computer_card_copy",
        "upload_trade_license_copy",
        "upload_vat_certificate_copy",
        "upload_food_safety_certificate",
        "upload_company_logo"
    ]

    for d in documents:
        docs.append([d, "Attach this document in email"])

    return wb


def send_excel_mail(to_email, workbook, filename):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    msg = MIMEMultipart()
    msg["Subject"] = "Action Required: Profile Completion Excel Template"

    sender_cfg = current_app.config["MAIL_DEFAULT_SENDER"]
    if isinstance(sender_cfg, tuple):
        name, email_addr = sender_cfg
        msg["From"] = f"{name} <{email_addr}>"
        from_addr = email_addr
    else:
        msg["From"] = sender_cfg
        from_addr = sender_cfg

    msg["To"] = to_email

    msg.attach(MIMEText(
        "Please fill the attached Excel template and attach required documents in reply to this email.",
        "plain"
    ))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(buffer.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={filename}")
    msg.attach(part)

    server = smtplib.SMTP(
        current_app.config["MAIL_SERVER"],
        current_app.config["MAIL_PORT"],
        timeout=20
    )

    try:
        server.connect(
            current_app.config["MAIL_SERVER"],
            current_app.config["MAIL_PORT"]
        )
        server.ehlo()
        server.starttls()
        server.ehlo()

        server.login(
            current_app.config["MAIL_USERNAME"],
            current_app.config["MAIL_PASSWORD"]
        )

        server.sendmail(from_addr, [to_email], msg.as_string())

    finally:
        server.quit()

@restapproval_bp.route("/restaurants/send-excel-template", methods=["POST"])
@require_admin("APPROVE_RESTAURANTS")
def send_excel_template():

    



    data = request.get_json() or {}
    entity_id = data.get("id")

    

    if not entity_id:
        return jsonify({"error": "restaurant id required"}), 400

    email = fetch_email("restaurant", entity_id)
    

    if not email:
        return jsonify({"error": "email not found"}), 400

    log_admin_action(
        admin_id=g.admin["admin_id"],
        action="SEND_EXCEL_TEMPLATE",
        entity_type="restaurant",
        entity_id=entity_id,
        new_value={"email": email},
        ip_address=request.remote_addr
    )

    workbook = build_restaurant_excel()
    filename = "restaurant_profile_template.xlsx"

    try:
        send_excel_mail(email, workbook, filename)
    except Exception as e:
        print("❌ Excel mail failed:", e)
        return jsonify({"error": "failed to send email"}), 500

    return jsonify({
        "status": True,
        "message": "Restaurant Excel template sent"
    })
# =========================================================
# 6️⃣ GET SUPPORT ADMINS (FOR ALLOTMENT)
# =========================================================
@restapproval_bp.route("/admins/support", methods=["GET", "OPTIONS"])
@require_admin("MANAGE_ADMIN_USERS")
def get_support_admins():

    if request.method == "OPTIONS":
        return "", 200

    conn = cur = None

    try:

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("""
            SELECT
                a.admin_id,
                a.name,

                COUNT(r.restaurant_id) AS pending

            FROM admin_users a

            JOIN admin_roles ar
                ON ar.role_id = a.role_id

            LEFT JOIN restaurant_registration r
                ON r.assigned_admin_id = a.admin_id
               AND r.approval_status = 'Pending'

            WHERE ar.role_name = 'SUPPORT_ADMIN'

            GROUP BY a.admin_id, a.name

            ORDER BY pending ASC, a.name ASC
        """)

        rows = cur.fetchall()

        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="READ_ACCESS",
            entity_type="support_admin_list",
            entity_id=0,
            ip_address=request.remote_addr
        )

        return jsonify({
            "admins": [
                {
                    "admin_id": r["admin_id"],
                    "name": r["name"],
                    "pending": r["pending"]
                }
                for r in rows
            ]
        })

    except Exception as e:

        print("❌ get_support_admins:", e)
        traceback.print_exc()

        return jsonify({"error": "server error"}), 500

    finally:

        if cur: cur.close()
        if conn: conn.close()

# =========================================================
# 7️⃣ ASSIGN / DEASSIGN RESTAURANT
# =========================================================
@restapproval_bp.route("/restaurant/<int:restaurant_id>/assign", methods=["PATCH"])
@require_admin("MANAGE_ADMIN_USERS")
def assign_restaurant(restaurant_id):

    data = request.get_json() or {}

    admin_id = data.get("admin_id")

    conn = cur = None

    try:

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT assigned_admin_id
            FROM restaurant_registration
            WHERE restaurant_id = %s
            FOR UPDATE
        """, (restaurant_id,))

        row = cur.fetchone()

        if not row:
            return jsonify({"error": "restaurant not found"}), 404

        old_admin = row["assigned_admin_id"]

        cur.execute("""
            UPDATE restaurant_registration
            SET assigned_admin_id = %s,
                updated_at = NOW()
            WHERE restaurant_id = %s
        """, (admin_id, restaurant_id))

        conn.commit()

        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="ASSIGN_RESTAURANT",
            entity_type="restaurant",
            entity_id=restaurant_id,
            old_value={"assigned_admin_id": old_admin},
            new_value={"assigned_admin_id": admin_id},
            ip_address=request.remote_addr
        )

        return jsonify({
            "status": True,
            "message": "Assignment updated"
        })

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ assign_restaurant:", e)
        traceback.print_exc()

        return jsonify({"error": "server error"}), 500

    finally:

        if cur: cur.close()
        if conn: conn.close()
# =========================================================
# 8️⃣ AUTO ASSIGN RESTAURANTS (LEAST PENDING FIRST)
# =========================================================
@restapproval_bp.route("/restaurants/auto-assign", methods=["PATCH"])
@require_admin("MANAGE_ADMIN_USERS")
def auto_assign_restaurants():

    conn = cur = None

    try:

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Get support admins ordered by least pending
        cur.execute("""
    SELECT
        a.admin_id,

        COUNT(r.restaurant_id) FILTER (
            WHERE r.approval_status = 'Pending'
        ) AS pending

    FROM admin_users a

    JOIN admin_roles ar
        ON ar.role_id = a.role_id

    LEFT JOIN restaurant_registration r
        ON r.assigned_admin_id = a.admin_id
       AND r.approval_status = 'Pending'

    WHERE ar.role_name = 'SUPPORT_ADMIN'

    GROUP BY a.admin_id

    ORDER BY pending ASC, a.admin_id ASC
""")


        admins = cur.fetchall()

        if not admins:
            return jsonify({"error": "no support admins"}), 400

        admin_ids = [a["admin_id"] for a in admins]

        # Get unassigned restaurants
        cur.execute("""
            SELECT restaurant_id
            FROM restaurant_registration
            WHERE approval_status = 'Pending'
              AND assigned_admin_id IS NULL
            ORDER BY created_at ASC
        """)

        restaurants = cur.fetchall()

        if not restaurants:
            return jsonify({
                "status": True,
                "message": "Nothing to assign"
            })

        index = 0

        for r in restaurants:

            admin_id = admin_ids[index]

            cur.execute("""
                UPDATE restaurant_registration
                SET assigned_admin_id = %s
                WHERE restaurant_id = %s
            """, (admin_id, r["restaurant_id"]))

            index = (index + 1) % len(admin_ids)

        conn.commit()

        log_admin_action(
            admin_id=g.admin["admin_id"],
            action="AUTO_ASSIGN_RESTAURANTS",
            entity_type="restaurant",
            entity_id=0,
            ip_address=request.remote_addr
        )

        return jsonify({
            "status": True,
            "message": "Auto assignment completed"
        })

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ auto_assign_restaurants:", e)
        traceback.print_exc()

        return jsonify({"error": "server error"}), 500

    finally:

        if cur: cur.close()
        if conn: conn.close()
